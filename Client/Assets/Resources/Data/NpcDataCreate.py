import openpyxl
import json

# 엑셀 파일 로드
workbook = openpyxl.load_workbook("NpcDataExel.xlsm")

# 각 탭에 대한 데이터 로드
npcs_sheet = workbook["npcs"]
npc_chats_sheet = workbook["npc_chats"]
npc_products_sheet = workbook["npc_products"]

# 결과 저장을 위한 dict
result_data = {"npcs": []}

# npcs 데이터 로드
for row in npcs_sheet.iter_rows(min_row=2, values_only=True):  # 첫 번째 row는 header 이므로 제외
    npc = {
        "npcId": str(row[0]),
        "name": row[1],
        "map": str(row[2]),
        "quest": str(row[3]),
        "posX": str(row[4]),
        "posY": str(row[5]),
        "iconPath": row[6],
        "merchant": str(row[7]),
        "chats": [],
        "products": []
    }
    result_data["npcs"].append(npc)

# npc_chats 데이터 로드
for row in npc_chats_sheet.iter_rows(min_row=2, values_only=True):
    for npc in result_data["npcs"]:
        if npc["npcId"] == str(row[0]):
            npc["chats"].append({
                "index": str(row[1]),
                "chat": row[2]
            })

# npc_products 데이터 로드
for row in npc_products_sheet.iter_rows(min_row=2, values_only=True):
    for npc in result_data["npcs"]:
        if npc["npcId"] == str(row[0]):
            npc["products"].append({"templateId": str(row[1])})

# 결과 데이터를 json 파일로 저장
with open("NpcData.json", "w", encoding="utf-8") as json_file:
    json.dump(result_data, json_file, ensure_ascii=False, indent=4)